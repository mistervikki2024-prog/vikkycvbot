from flask import Flask
import os
import threading
import json
import re
import time
import psutil
import telebot
from telebot import types
from threading import Lock
from zoneinfo import ZoneInfo
from datetime import datetime
import traceback


def safe_execute(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            print(f"\n❌ ERROR in {func.__name__}")
            traceback.print_exc()
    return wrapper

START_TIME = time.time()

def get_indian_time():
    return datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%d %b %Y, %I:%M:%S %p")

def get_uptime():
    seconds = int(time.time() - START_TIME)
    days = seconds // 86400
    hours = (seconds % 86400) // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    return f"{days}d {hours}h {minutes}m {seconds}s"

def build_stats_text(total_users, total_vcf):
    return f"""<b>📊 SYSTEM LIVE STATISTICS</b>
━━━━━━━━━━━━━━━━━━━━━━
<blockquote>
📈 <b>GLOBAL BOT USAGE</b>
├ 👥 <b>Total Users:</b> <code>{total_users}</code>
└ 📁 <b>VCFs Generated:</b> <code>{total_vcf}</code>
</blockquote><blockquote>
⚙️ <b>SERVER PERFORMANCE</b>
├ ⏱ <b>Uptime:</b> <code>{get_uptime()}</code>
├ 📡 <b>Ping Status:</b> (/ping)
├ 🎁 <b>Free Mode:</b> ON
└ 🟢 <b>Status:</b> Online
</blockquote><blockquote>
🤖 <b>SYSTEM HEALTH</b>
├ 🔥 <b>Load:</b> Optimal
├ ⚡ <b>Speed:</b> Fast Response
└ 🛡 <b>Security:</b> Protected
</blockquote>━━━━━━━━━━━━━━━━━━━━━━
👨‍💻 <b>Developed By:</b> <b>@Vikky_IND</b>  
🔄 <b>Last Updated:</b> <code>{get_indian_time()}</code>
"""


# ============================================================
# 🔹 ONLY VALID NUMBER EXTRACTION
# ============================================================
def extract_valid_numbers(text):
    numbers = re.findall(r'\d+', text)  # sirf digits nikalega
    valid = []

    for n in numbers:
        if len(n) >= 8:   # minimum length condition
            valid.append(n)

    return valid

msg_lock = Lock()
# 🔥 RATE LIMIT SYSTEM
last_used = {}

def is_rate_limited(user_id):
    now = time.time()
    actions = user_actions.get(user_id, [])

    # last 1 second ke actions
    actions = [t for t in actions if now - t < 1]

    if len(actions) >= 3:
        return True   # 1 sec me max 3 clicks

    actions.append(now)
    user_actions[user_id] = actions
    return False


# 🔹 Flask app
web = Flask(__name__)

@web.route('/')
def home():
    return "Bot is running!"

# 🔹 CONFIGRATION
TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID", "5328734113"))

bot = telebot.TeleBot(TOKEN)
user_state = {}
state_lock = Lock()
data_lock = Lock()


# ============================================================
# 🔹 GLOBAL DATA (STATS SYSTEM)
# ============================================================
DATA_FILE = "data.json"

# 🔹 USER SETTINGS (Caption Toggle)
SETTINGS_FILE = "settings.json"

def load_data():
    if not os.path.exists(DATA_FILE):
        return {"users": [], "vcf_count": 0}
    with open(DATA_FILE, "r") as f:
        return json.load(f)

def save_data(data):
    with data_lock:
        with open(DATA_FILE, "w") as f:
            json.dump(data, f, indent=4)


def load_settings():
    if not os.path.exists(SETTINGS_FILE):
        return {}
    with open(SETTINGS_FILE, "r") as f:
        return json.load(f)

def save_settings(data):
    with data_lock:
        with open(SETTINGS_FILE, "w") as f:
            json.dump(data, f, indent=4)

# ============================================================
# 🔹 MAIN MENU — Colored Buttons + Animated Emoji
# ============================================================
def main_menu():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    
    # Row 1
    kb.row(
        types.KeyboardButton("Text to VCF", style="primary", icon_custom_emoji_id="5433653135799228968"),
        types.KeyboardButton("VCF to Text", style="primary", icon_custom_emoji_id="5431736674147114227")
    )
    
    # Row 2
    kb.row(
        types.KeyboardButton("Admin/Navy VCF", style="danger", icon_custom_emoji_id="6266995104687330978"),
        types.KeyboardButton("Manual Text", style="danger", icon_custom_emoji_id="5334882760735598374")
    )
    
    # Row 3
    kb.row(
        types.KeyboardButton("Merge VCF", style="primary", icon_custom_emoji_id="5264727218734524899"),
        types.KeyboardButton("Merge Text", style="primary", icon_custom_emoji_id="5264727218734524899")
    )

    # Row 4
    kb.row(
        types.KeyboardButton("Split VCF", style="danger", icon_custom_emoji_id="5237808360882977239"),
        types.KeyboardButton("Split Text", style="danger", icon_custom_emoji_id="5237808360882977239")
    )
    
    # Row 5
    kb.row(
        types.KeyboardButton("VCF Editor", style="primary", icon_custom_emoji_id="5334673106202010226"),
        types.KeyboardButton("Get VCF Details", style="primary", icon_custom_emoji_id="5188217332748527444")
    )
    
    # Row 6
    kb.row(
        types.KeyboardButton("My Subscription", style="success", icon_custom_emoji_id="5422683699130933153")
    )
    
    return kb


# ============================================================
# 🔹 /start
# ============================================================
@bot.message_handler(commands=["start"])
@safe_execute
def start(message):
    uid = message.chat.id

#   user save for stats
    data = load_data()
    user_id = message.from_user.id
    if user_id not in data["users"]:
        data["users"].append(user_id)
        save_data(data)

    # 🔹 USER DATA
    user = message.from_user
    name = user.first_name
    username = f"@{user.username}" if user.username else "No Username"
    user_id = user.id

    # 🔥 animation me data pass kar
    if threading.active_count() < 40:
        threading.Thread(
            target=run_animation,
            args=(uid, name, username, user_id),
            daemon=True
            ).start()
    else:
        bot.send_message(uid, "⚠️ Server busy, please try again in a few seconds.")


@bot.message_handler(content_types=["photo"])
def get_file_id(message):
    file_id = message.photo[-1].file_id
    bot.reply_to(message, f"FILE_ID:\n{file_id}")

# ============================================================
# 🔹 STATS COMMAND (CLEAN VERSION)
# ============================================================
@bot.message_handler(commands=["stats"])
@safe_execute
def stats_cmd(message):

    data = load_data()

    total_users = len(set(data.get("users", [])))
    total_vcf = data.get("vcf_count", 0)

    text = build_stats_text(total_users, total_vcf)

    markup = types.InlineKeyboardMarkup()
    if message.from_user.id == ADMIN_ID:
        markup.add(
            types.InlineKeyboardButton("🔄 Refresh", callback_data="refresh_stats")
            )

    bot.send_message(
        message.chat.id,
        text,
        parse_mode="HTML",
        reply_markup=markup
    )


# ============================================================
# 🔹 STATS REFRESH BUTTON
# ============================================================
@bot.callback_query_handler(func=lambda call: call.data == "refresh_stats")
@safe_execute
def refresh_stats(call):
    if call.from_user.id != ADMIN_ID:
        bot.answer_callback_query(call.id, "❌ Only admin can refresh!", show_alert=True)
        return

    data = load_data()

    total_users = len(set(data.get("users", [])))
    total_vcf = data.get("vcf_count", 0)

    text = build_stats_text(total_users, total_vcf)

    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("🔄 Refresh", callback_data="refresh_stats")
    )

    bot.edit_message_text(
        text,
        call.message.chat.id,
        call.message.message_id,
        parse_mode="HTML",
        reply_markup=markup
    )

# ============================================================
# 🔹 /PING COMMAND
# ============================================================
@bot.message_handler(commands=["ping"])
@safe_execute
def ping_cmd(message):

    start = time.perf_counter()

    msg = bot.send_message(message.chat.id, "🏓 Checking...")

    ping = (time.perf_counter() - start) * 1000
    ping = round(ping, 2)

    uptime = get_uptime()

    cpu = psutil.cpu_percent()
    ram = psutil.virtual_memory().percent

    text = f"""<b>🏓 PONG! SYSTEM STATUS</b>
━━━━━━━━━━━━━━━━━━━━━━
<blockquote>
📡 <b>NETWORK LATENCY</b>
├⚡ <b>Speed:</b> Ultra Fast
└📶 <b>Ping:</b> <code>{ping} ms</code>
</blockquote><blockquote>
⚙️ <b>SYSTEM PERFORMANCE</b>
├⏱ <b>Uptime:</b> <code>{uptime}</code>
├🔥 <b>CPU Usage:</b> <code>{cpu}%</code>
└💾 <b>RAM Usage:</b> <code>{ram}%</code>
</blockquote><blockquote>
🟢 <b>SYSTEM STATUS</b>
├🛡 <b>Server:</b> Operational
├🚀 <b>Performance:</b> Stable & Smooth
└🤖 <b>Engine:</b> VCF Master Core
</blockquote>━━━━━━━━━━━━━━━━━━━━━━
👨‍💻 <b>Owner:</b> <b>@Vikky_IND</b>
🔄 <b>Last Updated:</b> <code>{get_indian_time()}</code>
"""

    markup = types.InlineKeyboardMarkup()
    if message.from_user.id == ADMIN_ID:
        markup.add(types.InlineKeyboardButton("🔄 Refresh", callback_data="refresh_ping"))

    bot.edit_message_text(
        text,
        message.chat.id,
        msg.message_id,
        parse_mode="HTML",
        reply_markup=markup
    )

# ============================================================
# 🔹 REFRESH PING
# ============================================================
@bot.callback_query_handler(func=lambda call: call.data == "refresh_ping")
def refresh_ping(call):
    if call.from_user.id != ADMIN_ID:
        bot.answer_callback_query(call.id, "❌ Only admin can refresh!", show_alert=True)
        return

    start = time.time()

    # 👉 temporary loading text
    bot.answer_callback_query(call.id, "Refreshing...")

    uptime = get_uptime()

    cpu = psutil.cpu_percent()
    ram = psutil.virtual_memory().percent

    # 👉 edit karte time actual delay aayega
    ping = int((time.time() - start) * 1000)

    text = f"""<b>🏓 PONG! SYSTEM STATUS</b>
━━━━━━━━━━━━━━━━━━━━━━
<blockquote>
📡 <b>NETWORK LATENCY</b>
├⚡ <b>Speed:</b> Ultra Fast
└📶 <b>Ping:</b> <code>{ping} ms</code>
</blockquote><blockquote>
⚙️ <b>SYSTEM PERFORMANCE</b>
├⏱ <b>Uptime:</b> <code>{uptime}</code>
├🔥 <b>CPU Usage:</b> <code>{cpu}%</code>
└💾 <b>RAM Usage:</b> <code>{ram}%</code>
</blockquote><blockquote>
🟢 <b>SYSTEM STATUS</b>
├🛡 <b>Server:</b> Operational
├🚀 <b>Performance:</b> Stable & Smooth
└🤖 <b>Engine:</b> VCF Master Core
</blockquote>━━━━━━━━━━━━━━━━━━━━━━
👨‍💻 <b>Owner:</b> <b>@Vikky_IND</b>
🔄 <b>Last Updated:</b> <code>{get_indian_time()}</code>
"""

    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("🔄 Refresh", callback_data="refresh_ping"))

    bot.edit_message_text(
        text,
        call.message.chat.id,
        call.message.message_id,
        parse_mode="HTML",
        reply_markup=markup
    )

@bot.message_handler(commands=["caption"])
@safe_execute
def caption_cmd(message):

    # 🔥 STEP 1: PHOTO (FIXED CAPTION - NOT EDITABLE)
    photo = "AgACAgUAAxkBAAIK1Gnop9zvKYD0WzJsbanCmvj9FgkKAAJxEGsbldpIV2RyntpKZm-yAQADAgADeQADOwQ"  # 👈 apni image

    bot.send_photo(
        message.chat.id,
        photo=photo,
        caption="You Want Caption Like This ?"
    )

    # 🔥 STEP 2: TEXT MESSAGE (EDITABLE)
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("✅ Yes", callback_data="cap_on"),
        types.InlineKeyboardButton("❌ No", callback_data="cap_off")
    )

    msg = bot.send_message(
        message.chat.id,
        "📸 Do you want a caption like the above on your VCF files?",
        reply_markup=markup
    )

    # 👉 message_id save karo (edit ke liye)
    with state_lock:
        user_state[message.from_user.id] = {"caption_msg_id": msg.message_id}

@bot.callback_query_handler(func=lambda call: call.data in ["cap_on", "cap_off"])
@safe_execute
def handle_caption_toggle(call):
    if is_rate_limited(user_id):
        bot.reply_to(message, "⚠️ Slow down! Too many requests.")
        return

    settings = load_settings()
    uid = str(call.from_user.id)

    if call.data == "cap_on":
        settings[uid] = {"caption": True}
        text = "✅ Caption preference set to ON."
    else:
        settings[uid] = {"caption": False}
        text = "❌ Caption turned OFF."

    save_settings(settings)

    # 🔥 ORIGINAL message id lo
    msg_id = user_state.get(call.from_user.id, {}).get("caption_msg_id")

    if msg_id:
        try:
            bot.edit_message_text(
                text,
                chat_id=call.message.chat.id,
                message_id=msg_id
            )
        except:
            pass

    bot.answer_callback_query(call.id)

# ============================================================
# 🔹 RUN ANIMATION
# ============================================================
def run_animation(uid, name, username, user_id):
    frames = [
        "[>_] INITIALIZING SYSTEM...\nEstablishing Secure Connection...\n🟥⬜️⬜️⬜️⬜️⬜️ 10%",
        "[>_] CONNECTING TO SERVERS...\nAuthorizing Access...\n🟥🟥⬜️⬜️⬜️⬜️ 30%",
        "[>_] BYPASSING FIREWALL...\nDecrypting Modules...\n🟥🟥🟥⬜️⬜️⬜️ 50%",
        "[>_] LOADING VCF ENGINE...\nOptimizing Performance...\n🟥🟥🟥🟥⬜️⬜️ 70%",
        "[>_] FINALIZING SETUP...\nLaunching Interface...\n🟥🟥🟥🟥🟥⬜️ 90%",
        "[✔] ACCESS GRANTED\nSYSTEM READY\n🟩🟩🟩🟩🟩🟩 100%"
    ]

    msg = bot.send_message(uid, f"<code>{frames[0]}</code>", parse_mode="HTML")

    for frame in frames[1:]:
        time.sleep(0.15)
        try:
            bot.edit_message_text(
                f"<code>{frame}</code>",
                chat_id=uid,
                message_id=msg.message_id,
                parse_mode="HTML"
            )
        except:
            pass
    try:
        bot.delete_message(uid, msg.message_id)
    except:
        pass

    # 🔥 FINAL PRO WELCOME (DYNAMIC)
    WELCOME_TEXT = f"""╔════════════════════════╗
    🔥 𝐖𝐄𝐋𝐂𝐎𝐌𝐄 𝐓𝐎 𝐕𝐂𝐅 𝐌𝐀𝐒𝐓𝐄𝐑 🔥
╚════════════════════════╝
<blockquote>👤 Name : {name}  
🔗 Username : {username}  
🆔 ID : {user_id}  
💎 Status : PREMIUM ACCESS 🔓  
</blockquote>
<blockquote>━━━━━━━━━━━━━━━━━━━━━━━
🛠️ BOT INFORMATION
━━━━━━━━━━━━━━━━━━━━━━━
🤖 System  : Advanced VCF Engine  
👨‍💻 Owner   : @Vikky_IND  
</blockquote>━━━━━━━━━━━━━━━━━━━━━━━
📩 Need help? Type → /help  
👇 Select a service from the menu below
"""

    bot.send_message(
    uid,
    WELCOME_TEXT,
    parse_mode="HTML",
    reply_markup=main_menu()
)


# ============================================================
# 🔹 User State
# ============================================================
def set_mode(user_id, mode):
    with state_lock:
        user_state[user_id] = {
            "mode": mode,
            "step": None,
            "data": {}
            }

# ============================================================
# 🔹 Load / Save Users
# ============================================================
def load_users():
    try:
        with open("users.json", "r") as f:
            return json.load(f)
    except:
        return {}

def save_users(data):
    with data_lock:
        with open("users.json", "w") as f:
            json.dump(data, f, indent=4)


# ============================================================
# 🔹 Progress Bar
# ============================================================
def progress_bar(current, total):
    percent = int((current / total) * 100) if total else 0
    filled = int(percent / 5)
    bar = "█" * filled + "░" * (20 - filled)
    return f"{bar} {percent}%"

def is_premium(user_id):
    data = load_data()
    return user_id in data.get("premium_users", [])

def add_premium(user_id):
    data = load_data()
    if "premium_users" not in data:
        data["premium_users"] = []

    if user_id not in data["premium_users"]:
        data["premium_users"].append(user_id)
        save_data(data)

def remove_premium(user_id):
    data = load_data()
    if user_id in data.get("premium_users", []):
        data["premium_users"].remove(user_id)
        save_data(data)

@bot.message_handler(commands=["addpremium"])
def give_premium(message):
    if message.from_user.id != ADMIN_ID:
        return

    try:
        uid = int(message.text.split()[1])
        add_premium(uid)
        bot.reply_to(message, "✅ User promoted to premium")
    except:
        bot.reply_to(message, "Usage: /addpremium user_id")

# ============================================================
# 🔹 HELP COMMAND
# ============================================================
@bot.message_handler(commands=["help"])
@safe_execute
def help_cmd(message):
    bot.send_message(
        message.chat.id,
        """
   🛠 HELP CENTER 🛠
━━━━━━━━━━━━━━━━━━━━━
🔥 𝐖𝐄𝐋𝐂𝐎𝐌𝐄 𝐓𝐎 𝐕𝐂𝐅 𝐌𝐀𝐒𝐓𝐄𝐑 🔥
Here is a quick guide to help you use all premium features efficiently:

👋 𝗚𝗘𝗧𝗧𝗜𝗡𝗚 𝗦𝗧𝗔𝗥𝗧𝗘𝗗
• /start → Start bot
• /done → Finish upload
• /cancel → Stop process

<blockquote>1️⃣ 𝗖𝗢𝗡𝗩𝗘𝗥𝗦𝗜𝗢𝗡 𝗧𝗢𝗢𝗟𝗦
━━━━━━━━━━━━━━━━━━━━━━━
➥ 📁 𝗧𝗲𝘅𝘁 𝘁𝗼 𝗩𝗖𝗙:- Send normal numbers, .txt, or .xlsx files and convert them into a ready-to-use VCF file.
➥ 🗂 𝗩𝗖𝗙 𝘁𝗼 𝗧𝗲𝘅𝘁:- Upload any VCF file to extract all contacts into a clean .txt file.
</blockquote>
<blockquote>2️⃣ 𝗩𝗖𝗙 𝗠𝗔𝗡𝗔𝗚𝗘𝗠𝗘𝗡𝗧
━━━━━━━━━━━━━━━━━━━━━━━
➥ 🔄 𝗠𝗲𝗿𝗴𝗲 𝗩𝗖𝗙:- Send multiple VCF files, and the bot will combine them into a single file.
➥ ✂️ 𝗦𝗽𝗹𝗶𝘁 𝗩𝗖𝗙:- Upload a large VCF file and split it into smaller parts (e.g., 50 contacts per file).
➥ ✏️ 𝗩𝗖𝗙 𝗘𝗱𝗶𝘁𝗼𝗿:- Upload existing VCF files, apply a new name/prefix, and export them instantly.
</blockquote>
<blockquote>3️⃣ 𝗦𝗣𝗘𝗖𝗜𝗔𝗟 𝗙𝗘𝗔𝗧𝗨𝗥𝗘𝗦
━━━━━━━━━━━━━━━━━━━━━━━
➥ 👑 𝗔𝗱𝗺𝗶𝗻 & 𝗡𝗮𝘃𝘆 𝗠𝗼𝗱𝗲:- Create segmented VCF files with different prefixes for Admin and Navy contacts automatically.
➥ 🔄 𝗠𝗲𝗿𝗴𝗲 𝗧𝗲𝘅𝘁:- Combine multiple .txt number files into a single file.
➥ 🔎 𝗩𝗖𝗙 𝗦𝗰𝗮𝗻𝗻𝗲𝗿:- Upload any VCF file to preview all names and numbers inside it.
➥ ✂️ 𝗦𝗽𝗹𝗶𝘁 𝗧𝗲𝘅𝘁:- Upload a large .txt file and split it into multiple smaller files for easier management.
</blockquote>
<blockquote>💡 𝗜𝗠𝗣𝗢𝗥𝗧𝗔𝗡𝗧 𝗣𝗥𝗢 𝗧𝗜𝗣𝗦
━━━━━━━━━━━━━━━━━━━━━━━
🔹 Always send /done after finishing file uploads or number input.

🔹 If you make a mistake, use /cancel to safely stop the process.
</blockquote>

<blockquote>👨‍💻 𝗢𝘄𝗻𝗲𝗿 & 𝗗𝗲𝘃𝗲𝗹𝗼𝗽𝗲𝗿:- @Vikky_IND
</blockquote>

""",
        parse_mode="HTML"
    )


# ============================================================
# 🔹 CANCEL COMMAND
# ============================================================
@bot.message_handler(commands=["cancel"])
@safe_execute
def cancel_cmd(message):
    user_id = message.from_user.id
    state = user_state.get(user_id)

    if state:
        state["cancelled"] = True  # 👈 STOP SIGNAL

        # 👉 update existing message if exists
        if state.get("msg_id"):
            try:
                bot.edit_message_text(
                    "❌ Process Cancelled!\n━━━━━━━━━━━━━━━\n🔄 You can start again.",
                    message.chat.id,
                    state["msg_id"]
                )
            except:
                pass

        # 👉 remove state
        user_state.pop(user_id, None)

    bot.send_message(
        message.chat.id,
        "❌ Process cancelled successfully.",
        reply_markup=main_menu()
    )



# ============================================================
# 🔹 TEXT HANDLER (FIXED)
# ============================================================
@bot.message_handler(func=lambda m: True, content_types=["text"])
@safe_execute
def handle_text(message):
    user_id = message.from_user.id
    if is_rate_limited(user_id):
        bot.reply_to(message, "⚠️ Slow down! Too many requests.")
        return
    text = message.text.strip()
    state = user_state.get(user_id)
    mode = state.get("mode") if state else None

    # ── MENU BUTTONS ──────────────────────────────────────────

    if text == "Text to VCF":
        start_txt_to_vcf(message, user_id)
        return

    if text == "VCF to Text":
        start_vcf_to_txt(message, user_id)
        return

    if text == "Admin/Navy VCF":
        start_admin_navy(message, user_id)
        return

    if text == "Manual Text":
        start_manual_text(message, user_id)
        return

    if text == "Merge VCF":
        start_merge_vcf(message, user_id)
        return

    if text == "Merge Text":
        start_merge_text(message, user_id)
        return

    if text == "Split VCF":
        start_split_vcf(message, user_id)
        return

    if text == "Split Text":
        start_split_text(message, user_id)
        return

    if text == "VCF Editor":
        start_vcf_editor(message, user_id)
        return

    if text == "Get VCF Details":
        start_vcf_details(message, user_id)
        return

    if text == "My Subscription":
        bot.send_message(
            message.chat.id,
            """🎉<b>BIG UPDATE: NOW FREE!</b> 🎉
<blockquote>━━━━━━━━━━━━━━━━━━━━━━━━
<b>All VIP & Premium features are unlocked for everyone — no subscription needed.</b>

🚀 <b>Enjoy full access to the bot’s advanced tools without limits!</b>
━━━━━━━━━━━━━━━━━━━━━━━━</blockquote>""",
            parse_mode="HTML",
            reply_markup=main_menu()
        )
        return

    # ── STATE CHECK ───────────────────────────────────────────

    if not state:
        bot.send_message(message.chat.id, "⚠️ Please select an option from menu first.", reply_markup=main_menu())
        return

    if mode == "admin_navy":
        handle_admin_navy(message, state, user_id)
        return

    if mode == "manual_text":
        handle_manual_text(message, state, user_id)
        return

    if mode == "split_vcf":
        handle_split_vcf(message, state, user_id)
        return

    if mode == "split_text":
        handle_split_text(message, state, user_id)
        return

# ── VCF TO TXT DONE ────────────────────────────────────
    if mode == "vcf_to_txt" and text == "/done":

        if not state["numbers"]:
            bot.send_message(message.chat.id, "❌ No data found.")
            return

        final_text = (
            f"📄 Extracted Numbers\n━━━━━━━━━━━━━━━\n"
            f"📁 Files Processed: {state.get('files', 0)}\n"
            f"📊 Final Extracted: {len(state['numbers'])}\n"
            f"✅ Finished!"
        )

        if state.get("msg_id"):
            try:
                bot.edit_message_text(final_text, message.chat.id, state["msg_id"])
            except:
                pass

        state["step"] = "ask_name"

        bot.send_message(
            message.chat.id,
            "📝 Enter the name for your .txt file:\nExample: ExtractedList"
        )
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("✅ Yes", callback_data="cap_on"),
            types.InlineKeyboardButton("❌ No", callback_data="cap_off")
        )

        bot.send_message(
            message.chat.id,
            "📸 Do you want caption on your VCF files?",
            reply_markup=markup
        )
        return

# ── TEXT TO VCF ────────────────────────────────────────────
    if mode == "txt_to_vcf":
        if state.get("step") == "collecting":
            handle_txt_input(message, state)
            return
        else:
            handle_txt_steps(message, state, user_id)
            return

    # ✅ ONLY EDIT — NO NEW MESSAGE
        if state.get("msg_id"):
            try:
                bot.edit_message_text(
                    final_text,
                    message.chat.id,
                    state["msg_id"]
                )
            except:
                pass

        state["step"] = "ask_name"

        bot.send_message(
            message.chat.id,
            "📝 Enter VCF file name:\nExample: Contacts"
        )
        return

    # 👉 FILE NAME INPUT
    if mode == "vcf_to_txt" and state.get("step") == "ask_name":
        filename = f"{text}.txt"

        with open(filename, "w") as f:
            f.write("\n".join(state["numbers"]))

        with open(filename, "rb") as f:
            bot.send_document(
                message.chat.id,
                f,
                caption="✅ Extracted Numbers"
            )

        os.remove(filename)

        bot.send_message(message.chat.id, "✅ Extraction Completed Successfully! 🎉")
        user_state.pop(user_id, None)
        return

# ── MERGE VCF DONE ─────────────────────────────
    if mode == "merge_vcf" and text == "/done":

        if not state["numbers"]:
            bot.send_message(message.chat.id, "❌ No data found.")
            return

        final_text = (
            "🔄 Merging VCF Files\n"
            "━━━━━━━━━━━━━━━\n"
            f"📊 Final Uploaded: {state['files']}\n"
            "✅ Finished!"
        )

        if state.get("msg_id"):
            try:
                bot.edit_message_text(final_text, message.chat.id, state["msg_id"])
            except:
                pass

        state["step"] = "ask_name"

        bot.send_message(
            message.chat.id,
            "📝 Enter the name for merged .vcf file:"
        )
        return

# ── MERGE VCF FILE NAME ───────────────────────
    if mode == "merge_vcf" and state.get("step") == "ask_name":

        filename = f"{text}.vcf"

        unique_numbers = list(set(state["numbers"]))

        vcf_lines = []
        count = 1

        for num in unique_numbers:
            vcf_lines.append(
                "BEGIN:VCARD\n"
                "VERSION:3.0\n"
                f"FN:Contact {count}\n"
                f"TEL;TYPE=CELL:{num}\n"
                "END:VCARD\n"
            )
            count += 1

        with open(filename, "w") as f:
            f.write("".join(vcf_lines))

        with open(filename, "rb") as f:
            bot.send_document(
                message.chat.id,
                f,
                caption="✅ Merged VCF"
            )
            data = load_data()
            data["vcf_count"] += 1
            save_data(data)

        os.remove(filename)

        bot.send_message(
            message.chat.id,
            "✅ Merging Completed Successfully! 🎉"
        )

        user_state.pop(user_id, None)
        return

# ── MERGE TEXT DONE ─────────────────────────────
    if mode == "merge_text" and text == "/done":

        if not state["lines"]:
            bot.send_message(message.chat.id, "❌ No data found.")
            return

        final_text = (
            "🔄 Merging Text Files\n"
            "━━━━━━━━━━━━━━━\n"
            f"📊 Final Uploaded: {state['files']}\n"
            "✅ Finished!"
        )

        if state.get("msg_id"):
            try:
                bot.edit_message_text(final_text, message.chat.id, state["msg_id"])
            except:
                pass

        state["step"] = "ask_name"

        bot.send_message(
            message.chat.id,
            "📝 Enter the name for merged .txt file:"
        )
        return

# ── MERGE TEXT FILE NAME ───────────────────────
    if mode == "merge_text" and state.get("step") == "ask_name":

        filename = f"{text}.txt"

        unique_lines = list(set(state["lines"]))

        with open(filename, "w", encoding="utf-8") as f:
            f.write("\n".join(unique_lines))

        with open(filename, "rb") as f:
            bot.send_document(
                message.chat.id,
                f,
                caption="✅ Merged Text"
            )

        os.remove(filename)

        bot.send_message(
            message.chat.id,
            "✅ Text Merging Completed Successfully! 🎉"
        )

        user_state.pop(user_id, None)
        return

    # ── VCF EDITOR DONE ──
    if mode == "vcf_editor" and text == "/done":

        if not state["contacts"]:
            bot.send_message(message.chat.id, "❌ No data found.")
            return

        final_text = (
            f"✏️ VCF Editor Mode\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📁 Files Processed: {state['files']}\n"
            f"📊 Contacts Found: {len(state['contacts'])}\n"
            f"✅ Finished!"
        )

        if state.get("msg_id"):
            try:
                bot.edit_message_text(final_text, message.chat.id, state["msg_id"])
            except:
                pass

        state["step"] = "ask_prefix"

        bot.send_message(
            message.chat.id,
            "🆔 Step 1 • New Contact Prefix\n"
            "━━━━━━━━━━━━━━━\n"
            "✏️ Enter the name you want for these contacts.\n\n"
            "Example: Rule Test"
        )
        return

    if mode == "vcf_editor":

    # STEP 1 → PREFIX
        if state.get("step") == "ask_prefix":
            state["prefix"] = text
            state["step"] = "ask_start"

            bot.send_message(
                message.chat.id,
                "🔢 Step 2 • Starting Number\n"
                "━━━━━━━━━━━━━━━\n"
                "🔢 Where should the counting start?\n\n"
                "Example: 1"
            )
            return

    # STEP 2 → START NUMBER
        if state.get("step") == "ask_start":
            if not text.isdigit():
                bot.send_message(message.chat.id, "❌ Enter valid number")
                return

            state["start"] = int(text)
            state["step"] = "ask_filename"

            bot.send_message(
                message.chat.id,
                "📁 Step 3 • VCF Filename\n"
                "━━━━━━━━━━━━━━━\n"
                "📝 Enter the name for your final exported VCF file.\n\n"
                "Example: Marketing"
            )
            return

    # STEP 3 → GENERATE
        if state.get("step") == "ask_filename":

            state["file_name"] = text
            state["step"] = "processing"

            bot.send_message(
                message.chat.id,
                f"🚀 Editing VCF Files\n"
                f"━━━━━━━━━━━━━━━\n"
                f"📁 Files: {state['files']}\n"
                f"👥 Total Contacts: {len(state['contacts'])}\n"
                f"⚡ Status: Processing..."
            )

            generate_edited_vcf(message, state, user_id)
            return

# ============================================================
# 🔹 START TXT TO VCF
# ============================================================
def start_txt_to_vcf(message, user_id):
    user_state[user_id] = {
        "mode": "txt_to_vcf",
        "step": "collecting",
        "numbers": [],
        "msg_id": None,
        "cancelled": False
    }

    bot.send_message(
        message.chat.id,
        "📥 Send Contacts\n━━━━━━━━━━━━━━━\n📂 Numbers / .txt / .xlsx\n\n✅ Finish Type → /done"
    )

# ============================================================
# 🔹 START VCF TO TXT
# ============================================================
def start_vcf_to_txt(message, user_id):
    user_state[user_id] = {
        "mode": "vcf_to_txt",
        "numbers": [],
        "files": 0,
        "msg_id": None,
        "cancelled": False
    }

    bot.send_message(
        message.chat.id,
        "📤 Upload VCF Files\n━━━━━━━━━━━━━━━\n📁 Send one or multiple .vcf files\n\n✅ Finish Type → /done"
    )

# ============================================================
# 🔹 START ADMIN\NAVY VCF
# ============================================================
def start_admin_navy(message, user_id):
    user_state[user_id] = {
        "mode": "admin_navy",
        "step": "admin_collect",
        "admin": [],
        "navy": [],
        "msg_id": None
    }

    text = (
        "1️⃣ Step 1 • Admin Contacts\n"
        "━━━━━━━━━━━━━━━\n"
        "📂 Send numbers or files\n\n"
        "⏭ Skip → /skip\n"
        "✅ Finish Type → /done"
    )

    msg = bot.send_message(message.chat.id, text)
    user_state[user_id]["msg_id"] = msg.message_id

# ============================================================
# 🔹 START MANUAL TEXT
# ============================================================
def start_manual_text(message, user_id):
    user_state[user_id] = {
        "mode": "manual_text",
        "step": "collect",
        "numbers": [],
        "msg_id": None
    }

    bot.send_message(
        message.chat.id,
        "📤 Send Contacts\n"
        "━━━━━━━━━━━━━━━\n"
        "📁 Send one or multiple numbers\n\n"
        "✅ Finish Type → /done"
    )

# ============================================================
# 🔹 START MERGE VCF
# ============================================================
def start_merge_vcf(message, user_id):
    user_state[user_id] = {
        "mode": "merge_vcf",
        "step": "collecting",
        "numbers": [],
        "files": 0,
        "msg_id": None
    }

    bot.send_message(
        message.chat.id,
        "🔄 Merge VCF Files\n━━━━━━━━━━━━━━━\n📁 Upload multiple .vcf files\n\n✅ Finish Type → /done"
    )

# ============================================================
# 🔹 START MERGE TEXT
# ============================================================
def start_merge_text(message, user_id):
    user_state[user_id] = {
        "mode": "merge_text",
        "step": "collecting",
        "lines": [],
        "files": 0,
        "msg_id": None
    }

    bot.send_message(
        message.chat.id,
        "🔄 Merge Text Files\n━━━━━━━━━━━━━━━\n📁 Upload multiple .txt files\n\n✅ Finish Type → /done"
    )

# ============================================================
# 🔹 START SPLIT VCF
# ============================================================
def start_split_vcf(message, user_id):
    user_state[user_id] = {
        "mode": "split_vcf",
        "step": "waiting_file",
        "file_path": None,
        "contacts": [],
        "file_name": None,
        "msg_id": None
    }

    bot.send_message(
        message.chat.id,
        "✂️ Split VCF File\n"
        "━━━━━━━━━━━━━━━\n"
        "📤 Send your large VCF file to split"
    )

# ============================================================
# 🔹 START SPLIT TEXT
# ============================================================
def start_split_text(message, user_id):
    user_state[user_id] = {
        "mode": "split_text",
        "step": "waiting_file",
        "lines": [],
        "file_path": None,
        "file_name": None
    }

    bot.send_message(
        message.chat.id,
        "✂️ Split Text File\n"
        "━━━━━━━━━━━━━━━\n"
        "📁 Upload ONE large .txt file to split"
    )

# ============================================================
# 🔹 START VCF EDITOR
# ============================================================
def start_vcf_editor(message, user_id):
    user_state[user_id] = {
        "mode": "vcf_editor",
        "step": "collecting",
        "contacts": [],
        "files": 0,
        "msg_id": None
    }

    bot.send_message(
        message.chat.id,
        "✏️ Upload VCF Files to Edit\n"
        "━━━━━━━━━━━━━━━\n"
        "📁 Send .vcf file(s)\n\n"
        "✅ Finish Type → /done"
    )

# ============================================================
# 🔹 START GET VCF DETAILS
# ============================================================
def start_vcf_details(message, user_id):
    user_state[user_id] = {
        "mode": "vcf_details",
        "page": 1,
        "contacts": [],
        "msg_id": None
    }

    msg = bot.send_message(
        message.chat.id,
        "📤 Upload a VCF file to see its details:"
    )

    user_state[user_id]["msg_id"] = msg.message_id

# ============================================================
# 🔹 UPDATE PROGRESS MESSAGE FOR TXT TO VCF
# ============================================================
def update_progress_message(message, state):
    msg_text = (
        f"📥 Collecting Contacts\n━━━━━━━━━━━━━━━\n"
        f"📊 Total Added: {len(state['numbers'])}\n"
        f"⏳ Status: Processing...\n\n"
        f"📂 Keep sending files/numbers\n"
        f"✅ Finish Type → /done"
    )

    with msg_lock:  # 🔴 LOCK START

        if not state.get("msg_id"):
            msg = bot.send_message(message.chat.id, msg_text)
            state["msg_id"] = msg.message_id

        else:
            try:
                bot.edit_message_text(
                    msg_text,
                    message.chat.id,
                    state["msg_id"]
                )
            except:
                # fallback (rare case)
                msg = bot.send_message(message.chat.id, msg_text)
                state["msg_id"] = msg.message_id

# ============================================================
# 🔹 UPDATE PROGRESS MESSAGE FOR VCF TO TXT
# ============================================================
def update_vcf_progress(message, state):
    msg_text = (
        f"📄 Extracting Numbers\n━━━━━━━━━━━━━━━\n"
        f"📁 Files Uploaded: {state['files']}\n"
        f"📊 Extracted: {len(state['numbers'])}\n"
        f"⏳ Status: Scanning...\n\n"
        f"📂 Keep sending files\n"
        f"✅ Finish Type → /done"
    )

    with msg_lock:
        if not state.get("msg_id"):
            msg = bot.send_message(message.chat.id, msg_text)
            state["msg_id"] = msg.message_id
        else:
            try:
                bot.edit_message_text(msg_text, message.chat.id, state["msg_id"])
            except:
                pass

# ============================================================
# 🔹 UPDATE PROGRESS MESSAGE FOR MERGE VCF
# ============================================================
def update_merge_progress(message, state):
    msg_text = (
        "🔄 Merging VCF Files\n"
        "━━━━━━━━━━━━━━━\n"
        f"📊 Uploaded VCFs: {state['files']}\n"
        "⏳ Status: Processing...\n\n"
        "📂 Keep sending files\n"
        "✅ Finish Type → /done"
    )

    with msg_lock:
        if not state.get("msg_id"):
            msg = bot.send_message(message.chat.id, msg_text)
            state["msg_id"] = msg.message_id
        else:
            try:
                bot.edit_message_text(msg_text, message.chat.id, state["msg_id"])
            except:
                pass

# ============================================================
# 🔹 UPDATE PROGRESS MESSAGE FOR MERGE TEXT
# ============================================================
def update_merge_text_progress(message, state):
    msg_text = (
        "🔄 Merging Text Files\n"
        "━━━━━━━━━━━━━━━\n"
        f"📊 Uploaded Files: {state['files']}\n"
        "⏳ Status: Processing...\n\n"
        "📂 Keep sending files\n"
        "✅ Finish Type → /done"
    )

    with msg_lock:
        if not state.get("msg_id"):
            msg = bot.send_message(message.chat.id, msg_text)
            state["msg_id"] = msg.message_id
        else:
            try:
                bot.edit_message_text(msg_text, message.chat.id, state["msg_id"])
            except:
                pass


# ============================================================
# 🔹 UPDATE MESSAGE FOR ADMIN NAVY VCF
# ============================================================
def update_admin_navy_msg(message, state, type_):

    if type_ == "admin":
        count = len(state["admin"])
        title = "👑 Step 1 • Collecting Admin"
        label = "Admin Added"
    else:
        count = len(state["navy"])
        title = "⚓ Step 2 • Collecting Navy"
        label = "Navy Added"

    text = (
        f"{title}\n"
        f"━━━━━━━━━━━━━━━\n"
        f"📊 {label}: {count}\n\n"
        f"📂 Keep sending files\n"
        f"✅ Finish → /done"
    )

    # ✅ FIRST TIME → NEW MESSAGE
    if not state.get("msg_id"):
        msg = bot.send_message(message.chat.id, text)
        state["msg_id"] = msg.message_id

    # ✅ NEXT TIME → EDIT SAME MESSAGE
    else:
        try:
            bot.edit_message_text(text, message.chat.id, state["msg_id"])
        except:
            pass



# ============================================================
# 🔹 HANDLE TEXT (TXT TO VCF FLOW)
# ============================================================
def handle_txt_input(message, state):
    if state.get("cancelled"):
        return
    text = message.text.strip()

    if text == "/done":
        if not state["numbers"]:
            bot.send_message(message.chat.id, "❌ No contacts added yet.")
            return

        final_text = (
            f"📥 Collected Contacts\n━━━━━━━━━━━━━━━\n"
            f"📊 Final Added: {len(state['numbers'])}\n"
            f"✅ Finished!"
        )

        if state.get("msg_id"):
            try:
                bot.edit_message_text(final_text, message.chat.id, state["msg_id"])
            except:
                pass

        state["step"] = "ask_file_name"
        bot.send_message(message.chat.id, "1️⃣ VCF File Name?\n(Example: HongKong)")
        return

    # 👉 ONLY ONE LOOP
    added = 0
    for n in text.split():
        n = n.replace("+","").replace("-","").replace(" ","")
        if n.isdigit() and len(n) >= 8:
            state["numbers"].append(n)
            added += 1

    if added > 0:
        update_progress_message(message, state)


# ============================================================
# 🔹 STEP FLOW (AFTER /done)
# ============================================================
def handle_txt_steps(message, state, user_id):
    text = message.text.strip()

    # 1️⃣ FILE NAME
    if state["step"] == "ask_file_name":
        state["file_name"] = text
        state["step"] = "ask_prefix"
        bot.send_message(message.chat.id, "2️⃣ Contact Name Prefix?\n(Example: Vikky Boss)")
        return

    # 2️⃣ PREFIX
    if state["step"] == "ask_prefix":
        state["prefix"] = text
        state["step"] = "ask_vcf_start"
        bot.send_message(message.chat.id, "3️⃣ VCF File Starting Number?\n(Example: 1)")
        return

    # 3️⃣ VCF START
    if state["step"] == "ask_vcf_start":
        if not text.isdigit():
            bot.send_message(message.chat.id, "❌ Enter valid number")
            return
        state["vcf_start"] = int(text)
        state["step"] = "ask_contact_start"
        bot.send_message(message.chat.id, "4️⃣ Contact Starting Number?\n(Example: 1)")
        return

    # 4️⃣ CONTACT START
    if state["step"] == "ask_contact_start":
        if not text.isdigit():
            bot.send_message(message.chat.id, "❌ Enter valid number")
            return
        state["contact_start"] = int(text)
        state["step"] = "ask_limit"
        bot.send_message(message.chat.id, "5️⃣ Contacts per VCF file?\n(Example: 50)")
        return

    # 5️⃣ LIMIT → GENERATE
    if state["step"] == "ask_limit":
        if not text.isdigit():
            bot.send_message(message.chat.id, "❌ Enter valid number")
            return

        limit = int(text)

    # 🔥 LIMIT SAFETY
        if limit > 500:
            bot.send_message(message.chat.id, "⚠️ Max limit is 500 per file. Auto set to 500.")
            limit = 500

    generate_vcf_files_clean(message, state, user_id, limit)

# ============================================================
# 🔹 CLEAN VCF GENERATOR (NO BUG)
# ============================================================
def generate_vcf_files_clean(message, state, user_id, limit):
    numbers = state["numbers"]

    bot.send_message(
        message.chat.id,
        f"🚀 Generating VCF Files\n━━━━━━━━━━━━━━━\n"
        f"📊 Total Contacts: {len(numbers)}\n"
        f"⚡ Status: Processing..."
    )

    file_index = state["vcf_start"]
    contact_counter = state["contact_start"]

    total = len(numbers)

    for i in range(0, total, limit):
        if state.get("cancelled"):
            bot.send_message(message.chat.id,"Process Stopped.")
            return
        chunk = numbers[i:i+limit]

        # ⚡ FAST BUILD (list + join)
        vcf_lines = []
        for num in chunk:
            vcf_lines.append(
                "BEGIN:VCARD\n"
                "VERSION:3.0\n"
                f"FN:{state['prefix']} {contact_counter}\n"
                f"TEL;TYPE=CELL:{num}\n"
                "END:VCARD\n"
            )
            contact_counter += 1

        vcf_data = "".join(vcf_lines)

        filename = f"{state['file_name']}{file_index}.vcf"
        file_index += 1

        # ⚡ FAST WRITE
        with open(filename, "w", encoding="utf-8") as f:
            f.write(vcf_data)

        # ⚡ SEND FILE
        settings = load_settings()
        uid = str(message.from_user.id)
        use_caption = settings.get(uid, {}).get("caption", False)
        numbers_list = state.get("numbers", [])
        total_contacts = len(numbers_list)
        contact_name = state.get("navy_name") or state.get("prefix") or "Contacts" if state else "Contacts"
        caption = f"""
├▸📁 <code>{filename}</code>
├▸👤 <code>{contact_name}</code>
└▸📊 <code>{len(chunk)}</code> <code>contacts</code>"""
        if use_caption:
            bot.send_document(
                message.chat.id,
                open(filename, "rb"),
                caption=caption,
                parse_mode="HTML"
                )
        else:
            bot.send_document(
                message.chat.id,
                open(filename, "rb")
            )
            data = load_data()
            data["vcf_count"] += 1
            save_data(data)

        os.remove(filename)

    bot.send_message(message.chat.id, "✅ VCF Generation Completed Successfully! 🎉")
    user_state.pop(user_id, None)

# ============================================================
# 🔹 HANDLE ADMIN NAVY
# ============================================================
def handle_admin_navy(message, state, user_id):
    text = message.text.strip()
    # 🔥 SKIP BUTTON SUPPORT
    if text.lower() in ["/skip", "skip"]:

        # 👉 STEP 1 SKIP (ADMIN → NAVY)
        if state["step"] == "admin_collect":

            final = (
                "👑 Step 1 • Admin Contacts\n"
                "━━━━━━━━━━━━━━━\n"
                "⏭ Skipped!\n"
                "📊 Final Admin: 0\n"
            )

            bot.edit_message_text(final, message.chat.id, state["msg_id"])

            state["step"] = "navy_collect"

            msg = bot.send_message(
                message.chat.id,
                "2️⃣ Step 2 • Navy Contacts\n"
                "━━━━━━━━━━━━━━━\n"
                "📂 Send Navy numbers or files\n\n"
                "⏭ Skip → /skip\n"
                "✅ Finish → /done"
            )

            state["msg_id"] = msg.message_id
            return

        # 👉 STEP 2 SKIP (NAVY → NEXT STEP)
        elif state["step"] == "navy_collect":

            final = (
                "⚓ Step 2 • Navy Contacts\n"
                "━━━━━━━━━━━━━━━\n"
                "⏭ Skipped!\n"
                "📊 Final Navy: 0\n"
            )

            bot.edit_message_text(final, message.chat.id, state["msg_id"])

            state["step"] = "ask_admin_name"

            bot.send_message(
                message.chat.id,
                "🖋 Step 3 • Admin Name Prefix\n"
                "━━━━━━━━━━━━━━━\n"
                "✏️ Enter Admin contact name.\n\n"
                "Example: Admin Target"
            )
            return

    # STEP 1 → ADMIN COLLECT
    if state["step"] == "admin_collect":

        if text == "/done":
            final = (
                "👑 Step 1 • Admin Contacts\n"
                "━━━━━━━━━━━━━━━\n"
                f"📊 Final Admin: {len(state['admin'])}\n"
                "✅ Saved!"
            )

    # ✅ SAME MESSAGE EDIT
            bot.edit_message_text(final, message.chat.id, state["msg_id"])

    # ✅ STEP CHANGE
            state["step"] = "navy_collect"

    # ✅ NEW MESSAGE (NAVY START)
            msg = bot.send_message(
                message.chat.id,
                "2️⃣ Step 2 • Navy Contacts\n"
                "━━━━━━━━━━━━━━━\n"
                "📂 Send Navy numbers or files.\n\n"
                "⏭ Skip → skip\n"
                "✅ Finish → /done"
            )

            state["msg_id"] = msg.message_id
            return

        # ADD NUMBERS
        nums = extract_valid_numbers(text)
        state["admin"].extend(nums)

        update_admin_navy_msg(message, state, "admin")
        return

    # STEP 2 → NAVY
    if state["step"] == "navy_collect":

        if text == "/done":
            final = (
                "⚓ Step 2 • Navy Contacts\n"
                "━━━━━━━━━━━━━━━\n"
                f"📊 Final Navy: {len(state['navy'])}\n"
                "✅ Saved!"
            )

    # ✅ SAME MESSAGE EDIT
            bot.edit_message_text(final, message.chat.id, state["msg_id"])

    # ✅ STEP CHANGE
            state["step"] = "ask_admin_name"

    # ✅ NEXT STEP MESSAGE
            bot.send_message(
                message.chat.id,
                "🖋 Step 3 • Admin Name Prefix\n"
                "━━━━━━━━━━━━━━━\n"
                "✏️ What should be the name for Admin contacts?\n\n"
                "Example: Admin Target"
            )
            return

        nums = extract_valid_numbers(text)
        state["navy"].extend(nums)

        update_admin_navy_msg(message, state, "navy")
        return

    # STEP 3
    if state["step"] == "ask_admin_name":
        state["admin_name"] = message.text.strip()
        state["step"] = "ask_navy_name"

        bot.send_message(
            message.chat.id,
            "🖋 Step 4 • Navy Name Prefix\n"
            "━━━━━━━━━━━━━━━\n"
            "✏️ Enter the name for Navy contacts.\n\n"
            "Example: Navy Target"
        )
        return

    # STEP 4
    if state["step"] == "ask_navy_name":
        state["navy_name"] = message.text.strip()
        state["step"] = "ask_admin_start"

        bot.send_message(
            message.chat.id,
            "🔢 Step 5 • Admin Start Number\n"
            "━━━━━━━━━━━━━━━\n"
            "🔢 Send start number for Admin contacts.\n\n"
            "⏭ Skip → /skip (Default: 1)"
        )
        return

    # STEP 5
    if state["step"] == "ask_admin_start":
        state["admin_start"] = int(text) if text.isdigit() else 1
        state["step"] = "ask_navy_start"

        bot.send_message(
            message.chat.id,
            "🔢 Step 6 • Navy Start Number\n"
            "━━━━━━━━━━━━━━━\n"
            "🔢 Send start number for Navy contacts.\n\n"
            "⏭ Skip → /skip (Default: 1)"
        )
        return

    # STEP 6
    if state["step"] == "ask_navy_start":
        state["navy_start"] = int(text) if text.isdigit() else 1
        state["step"] = "ask_filename"

        bot.send_message(
            message.chat.id,
            "📁 Step 7 • Final VCF Filename\n"
            "━━━━━━━━━━━━━━━\n"
            "📝 Enter the name for your generated VCF file.\n\n"
            "Example: Admin File"
        )
        return

    # STEP 7 → GENERATE
    if state["step"] == "ask_filename":
        filename = f"{text}.vcf"

        vcf = ""

        i = state["admin_start"]
        name = state["admin_name"]
        for num in state["admin"]:
            vcf += (
                "BEGIN:VCARD\n"
                "VERSION:3.0\n"
                f"FN:{name} {i}\n"
                f"TEL:{num}\n"
                "END:VCARD\n"
            )
            i += 1

        j = state["navy_start"]
        name = state["navy_name"]
        for num in state["navy"]:
            vcf += (
                "BEGIN:VCARD\n"
                "VERSION:3.0\n"
            f"FN:{name} {j}\n"
            f"TEL:{num}\n"
            "END:VCARD\n"
            )
            j += 1

        with open(filename, "w") as f:
            f.write(vcf)

        with open(filename, "rb") as f:
            bot.send_document(
                message.chat.id,
                f,
                caption="✅ Generated VCF"
            )
            data = load_data()
            data["vcf_count"] += 1
            save_data(data)

        os.remove(filename)

        bot.send_message(message.chat.id, "✅ Generation Completed! 🎉")
        user_state.pop(user_id, None)


# ============================================================
# 🔹 MANUAL TEXT
# ============================================================
def handle_manual_text(message, state, user_id):
    text = message.text.strip()

    # STEP 1 → COLLECT
    if state["step"] == "collect":

        if text == "/done":

            final_text = (
                "📄 Collected Numbers\n"
                "━━━━━━━━━━━━━━━\n"
                f"📊 Total added: {len(state['numbers'])}\n"
                "✅ Saved!"
            )

            if state.get("msg_id"):
                bot.edit_message_text(final_text, message.chat.id, state["msg_id"])

            state["step"] = "ask_name"

            bot.send_message(
                message.chat.id,
                "📝 Enter file name\nExample: MyList"
            )
            return

        # 🔥 BULK SAFE PARSER (FIXED)
        import re

        added = 0
        numbers = re.findall(r'\d+', text)
        for n in numbers:
            if len(n) >= 8:
                state["numbers"].append(n)
                added += 1

        if added == 0:
            return

        # ✅ SINGLE MESSAGE UPDATE (NO SPAM)
        msg_text = (
            "📄 Collecting Numbers\n"
            "━━━━━━━━━━━━━━━\n"
            "⏳ Status: Saving...\n"
            f"📊 Numbers added: {len(state['numbers'])}\n\n"
            "👤 Keep sending numbers\n"
            "✅ Finish Type → /done"
        )

        with msg_lock:
            if not state.get("msg_id"):
                msg = bot.send_message(message.chat.id, msg_text)
                state["msg_id"] = msg.message_id
            else:
                try:
                    bot.edit_message_text(msg_text, message.chat.id, state["msg_id"])
                except:
                    pass

    # STEP 2 → FILE NAME
    elif state["step"] == "ask_name":

        filename = f"{text}.txt"

        with open(filename, "w") as f:
            unique_numbers = list(set(state["numbers"]))
            f.write("\n".join(unique_numbers))

        with open(filename, "rb") as f:
            bot.send_document(message.chat.id, f,
            caption="✅ Generated Text File"
            )

        os.remove(filename)

        bot.send_message(message.chat.id, "✅ Text generated successfully")
        user_state.pop(user_id, None)

# ============================================================
# 🔹 HANDLE SPLIT VCF
# ============================================================
def handle_split_vcf(message, state, user_id):
    text = message.text.strip()

    # STEP 1 → LIMIT
    if state["step"] == "waiting_file":
        if not text.isdigit():
            return

        state["limit"] = int(text)
        state["step"] = "ask_name"

        kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
        kb.add(types.KeyboardButton("🔄 Same as Old"))

        old_name = state["file_path"].split("_",1)[-1].replace(".vcf","")

        bot.send_message(
            message.chat.id,
            f"📁 VCF File Name?\n"
            f"━━━━━━━━━━━━━━━\n"
            f"Type a new name OR click below to keep old name: {old_name}",
            reply_markup=kb
        )
        return

    # STEP 2 → FILE NAME
    if state["step"] == "ask_name":
        if text == "🔄 Same as Old":
            state["file_name"] = state["file_path"].split("_",1)[-1].replace(".vcf","")
        else:
            state["file_name"] = text

        state["step"] = "ask_prefix"

        kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
        kb.add(types.KeyboardButton("🔄 Same as Old"))

        bot.send_message(
            message.chat.id,
            "👤 Contact Name Prefix?\n"
            "━━━━━━━━━━━━━━━\n"
            "Type a new name OR click below to keep old names",
            reply_markup=kb
        )
        return

    # STEP 3 → PREFIX
    if state["step"] == "ask_prefix":
        if text == "🔄 Same as Old":
            state["prefix"] = None
        else:
            state["prefix"] = text

        state["step"] = "splitting"

        bot.send_message(
            message.chat.id,
            f"✂️ Splitting VCF Files...\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📊 Total Contacts: {len(state['contacts'])}\n"
            f"⚡ Status: Processing..."
        )

        split_vcf_files(message, state, user_id)

# ============================================================
# 🔹 SPILITING VCF AND SET FILE NAME 
# ============================================================
def split_vcf_files(message, state, user_id):
    contacts = state["contacts"]
    limit = state["limit"]
    filename = state["file_name"]
    prefix = state.get("prefix")

    file_index = 1

    for i in range(0, len(contacts), limit):
        chunk = contacts[i:i+limit]

        new_vcf = ""
        count = 1

        for v in chunk:
            if prefix:
                # replace FN line
                lines = v.split("\n")
                new_lines = []
                for line in lines:
                    if line.startswith("FN:"):
                        new_lines.append(f"FN:{prefix} {count}")
                    else:
                        new_lines.append(line)
                v = "\n".join(new_lines)

            new_vcf += v
            count += 1

        file_name = f"{filename}_{file_index}.vcf"
        file_index += 1

        with open(file_name, "w", encoding="utf-8") as f:
            f.write(new_vcf)

        with open(file_name, "rb") as f:
            bot.send_document(message.chat.id, f)
            data = load_data()
            data["vcf_count"] += 1
            save_data(data)

        os.remove(file_name)
        try:
            os.remove(state["file_path"])
        except:
            pass

    bot.send_message(
        message.chat.id,
        "✅ VCF Splitting Completed! 🎉",
        reply_markup=main_menu()
        )
    user_state.pop(user_id, None)


# ============================================================
# 🔹 HANDLE SPLIT TEXT
# ============================================================
def handle_split_text(message, state, user_id):
    text = message.text.strip()

    # STEP 1 → PARTS
    if state["step"] == "waiting_file":
        if not text.isdigit():
            return

        parts = int(text)
        total = len(state["lines"])

        if parts <= 0:
            return

        if parts > total:
            parts = total

        state["parts"] = parts
        state["step"] = "ask_name"

        bot.send_message(
            message.chat.id,
            "📁 Enter file name:\nExample: SplitFile"
        )
        return

    # STEP 2 → FILE NAME
    if state["step"] == "ask_name":
        if not text:
            text = "SplitFile"

        state["file_name"] = text
        state["step"] = "splitting"

        bot.send_message(
            message.chat.id,
            f"✂️ Splitting Text File...\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📊 Total Lines: {len(state['lines'])}\n"
            f"📂 Parts: {state['parts']}\n"
            f"⚡ Processing..."
        )

        split_text_files(message, state, user_id)

# ============================================================
# 🔹 SPILITING TEXT AND SET FILE NAME 
# ============================================================
def split_text_files(message, state, user_id):
    lines = state["lines"]
    parts = state["parts"]
    filename = state["file_name"]

    total = len(lines)

    per_file = total // parts
    extra = total % parts

    start = 0

    for i in range(parts):
        end = start + per_file + (1 if i < extra else 0)
        chunk = lines[start:end]

        file_name = f"{filename}_{i+1}.txt"

        with open(file_name, "w", encoding="utf-8") as f:
            f.write("\n".join(chunk))

        with open(file_name, "rb") as f:
            bot.send_document(message.chat.id, f)

        os.remove(file_name)

        start = end

    bot.send_message(
        message.chat.id,
        "✅ Text Splitting Completed! 🎉",
        reply_markup=main_menu()
    )

    user_state.pop(user_id, None)

# ============================================================
# 🔹 GENERATE EDITED VCF
# ============================================================
def generate_edited_vcf(message, state, user_id):
    contacts = state["contacts"]
    prefix = state["prefix"]
    start = state["start"]
    filename = state["file_name"]

    vcf_data = ""
    count = start

    for v in contacts:
        lines = v.split("\n")
        new_lines = []

        for line in lines:
            if line.startswith("FN:"):
                new_lines.append(f"FN:{prefix} {count}")
            else:
                new_lines.append(line)

        vcf_data += "\n".join(new_lines) + "\n"
        count += 1

    file_name = f"{filename}.vcf"

    with open(file_name, "w", encoding="utf-8") as f:
        f.write(vcf_data)

    with open(file_name, "rb") as f:
        bot.send_document(message.chat.id, f)
        data = load_data()
        data["vcf_count"] += 1
        save_data(data)

    os.remove(file_name)

    bot.send_message(
        message.chat.id,
        f"✅ Editing Completed Successfully! 🎉\n"
        f"━━━━━━━━━━━━━━━\n"
        f"📊 Final Contacts: {len(contacts)}\n"
        f"⚡ Files Generated: 1"
    )

    user_state.pop(user_id, None)

# ============================================================
# 🔹 GENERATE TXT REPORT
# ============================================================
def generate_txt_report(state):
    contacts = state["contacts"]
    filename = state.get("file_name", "Unknown.vcf")

    report = ""
    report += "=========================================\n"
    report += "          VCF DETAILS REPORT\n"
    report += "=========================================\n"
    report += f"File Name   : {filename}\n"
    report += f"Total       : {len(contacts)} Contacts\n"
    report += "Generated by : @Vikky_IND VCF Engine\n"
    report += "=========================================\n\n"

    for i, (name, phone) in enumerate(contacts, start=1):
        report += f"[{i}] Name : {name}\n"
        report += f"    Phone: {phone}\n\n"

    return report


# ============================================================
# 🔹 SEND TXT REPORT
# ============================================================
def send_txt_report(chat_id, state):
    report_text = generate_txt_report(state)

    # ✅ NAME FORMAT: originalname_report.txt
    original_name = state.get("file_name", "VCF_Report.vcf")
    base_name = original_name.rsplit(".", 1)[0]
    file_name = f"{base_name}_report.txt"

    with open(file_name, "w", encoding="utf-8") as f:
        f.write(report_text)

    with open(file_name, "rb") as f:
        bot.send_document(
            chat_id,
            f,
            caption="✅ Scanning Completed Successfully! 🎉\n📁 Full Report: 📥 Download .txt"
        )

    os.remove(file_name)

# ============================================================
# 🔹 Animate Progress
# ============================================================
def animate_progress(chat_id, msg_id, state):
    last_done = 0
    last_time = time.time()

    while state.get("animating"):
        time.sleep(0.5)
        total = max(state.get("total_lines", 1), 1)
        done = state.get("processed_lines", 0)

        now = time.time()
        speed = (done - last_done) / (now - last_time) if (now - last_time) > 0 else 0
        last_done = done
        last_time = now

        percent = min(int((done / total) * 100), 100)
        filled = int(percent / 5)
        bar = "█" * filled + "░" * (20 - filled)

        text_msg = (
            f"🚀 *VCF SCANNING*\n"
            f"━━━━━━━━━━━━━━━\n\n"
            f"📁 Files: {state.get('files', 0)}\n"
            f"📊 Extracted: {len(state.get('numbers', []))}\n\n"
            f"📈 Progress: `{bar} {percent}%`\n\n"
            f"⚡ Speed: {speed:.0f} lines/sec\n"
            f"🔄 {done}/{total} lines"
        )

        try:
            bot.edit_message_text(text_msg, chat_id, msg_id, parse_mode="Markdown")
        except:
            pass

# ============================================================
# 🔹 Process VCF File
# ============================================================
def process_vcf_file(path, state):
    with open(path, encoding="utf-8", errors="ignore") as f:
        for line in f:
            state["total_lines"] += 1
            line = line.strip()
            if "TEL" in line.upper():
                num = line.split(":")[-1].strip()
                num = num.replace(" ", "").replace("-", "").replace("+", "")
                if num.isdigit() and len(num) >= 8:
                    state["numbers"].append(num)
            state["processed_lines"] += 1
    try:
        os.remove(path)
    except:
        pass

# ============================================================
# 🔹 CALLBACK HANDLER
# ============================================================
@bot.callback_query_handler(func=lambda call: call.data in ["next", "prev", "refresh_stats", "refresh_ping"])
def callback_handler(call):
    user_id = call.from_user.id
    state = user_state.get(user_id)

    if not state or state.get("mode") != "vcf_details":
        return

    if call.data == "next":
        state["page"] += 1

    elif call.data == "prev":
        state["page"] -= 1

    show_vcf_page(call.message.chat.id, state)

# ============================================================
# 🔹 FILE HANDLER
# ============================================================
@bot.message_handler(content_types=["document"])
def handle_files(message):
    user_id = message.from_user.id
    state = user_state.get(user_id)

    if not state:
        bot.send_message(message.chat.id, "⚠️ Please select an option from menu first.")
        return

    if state.get("cancelled"):
        return

    mode = state.get("mode")
    doc = message.document
    filename = doc.file_name.lower()

    file_info = bot.get_file(doc.file_id)
    path = f"{user_id}_{filename}"

    downloaded = bot.download_file(file_info.file_path)
    with open(path, "wb") as f:
        f.write(downloaded)

    # ===== TXT =====
    if filename.endswith(".txt") and mode == "txt_to_vcf":
        with open(path) as f:
            for line in f:
                n = line.strip().replace("+","").replace("-","").replace(" ","")
                if n.isdigit() and len(n) >= 8:
                    state["numbers"].append(n)
        os.remove(path)

    # ===== XLSX =====
    elif filename.endswith(".xlsx") and mode == "txt_to_vcf":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        for row in wb.active.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    n = str(cell).strip().replace("+","").replace("-","").replace(" ","")
                    if n.isdigit() and len(n) >= 8:
                        state["numbers"].append(n)
        wb.close()
        os.remove(path)

    # ===== 🔥 UPDATE SAME MESSAGE =====
    if mode == "txt_to_vcf" and (filename.endswith(".txt") or filename.endswith(".xlsx")):
        update_progress_message(message, state)
        return

    # ===== VCF TO TXT =====
    elif filename.endswith(".vcf") and mode == "vcf_to_txt":
        state["files"] += 1

        with open(path, encoding="utf-8", errors="ignore") as f:
            for line in f:
                if "TEL" in line.upper():
                    num = line.split(":")[-1].strip()
                    num = num.replace(" ", "").replace("-", "").replace("+", "")
                    if num.isdigit() and len(num) >= 8:
                        state["numbers"].append(num)

        os.remove(path)

        update_vcf_progress(message, state)
        return

    # ===== MERGE VCF =====
    if filename.endswith(".vcf") and mode == "merge_vcf":
        state["files"] += 1

        with open(path, encoding="utf-8", errors="ignore") as f:
            for line in f:
                if "TEL" in line.upper():
                    num = line.split(":")[-1].strip()
                    num = num.replace(" ", "").replace("-", "").replace("+", "")
                    if num.isdigit() and len(num) >= 8:
                        state["numbers"].append(num)

        os.remove(path)

        update_merge_progress(message, state)
        return

    # ===== MERGE TEXT =====
    if filename.endswith(".txt") and mode == "merge_text":

        state["files"] += 1

        with open(path, encoding="utf-8", errors="ignore") as f:
            for line in f:
                clean = line.strip()
                if clean:
                    state["lines"].append(clean)

        os.remove(path)

        update_merge_text_progress(message, state)
        return

# ===== SPLIT VCF =====
    elif filename.endswith(".vcf") and mode == "split_vcf":

        contacts = []

        with open(path, encoding="utf-8", errors="ignore") as f:
            temp = ""
            for line in f:
                temp += line
                if "END:VCARD" in line:
                    contacts.append(temp)
                    temp = ""

        state["contacts"] = contacts
        state["file_path"] = path

        total = len(contacts)
        name = filename.replace(".vcf", "")

        bot.send_message(
            message.chat.id,
            f"✅ VCF Loaded!\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📁 File Name: {name}\n"
            f"👥 Total Contacts: {total}\n\n"
            f"🔢 How many contacts do you want per file? (e.g., 50, 100)"
        )
        return

    # ===== SPLIT TEXT =====
    elif filename.endswith(".txt") and mode == "split_text":

        with open(path, encoding="utf-8", errors="ignore") as f:
            lines = [line.strip() for line in f if line.strip()]

        state["lines"] = lines
        state["file_path"] = path

        total = len(lines)
        name = filename.replace(".txt", "")

        # 👉 auto filename set
        state["file_name"] = name

        bot.send_message(
            message.chat.id,
            f"✅ Text Loaded!\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📁 File Name: {name}\n"
            f"📄 Total Lines: {total}\n\n"
            f"🔢 How many parts do you want to divide?\n"
            f"(Example: 2 / 5 / 10)"
        )
        return

    # ===== VCF EDITOR =====
    elif filename.endswith(".vcf") and mode == "vcf_editor":
        state["files"] += 1

        contacts = []
        with open(path, encoding="utf-8", errors="ignore") as f:
            temp = ""
            for line in f:
                temp += line
                if "END:VCARD" in line:
                    contacts.append(temp)
                    temp = ""

        state["contacts"].extend(contacts)

        os.remove(path)

    # 🔥 MESSAGE UPDATE (LOCK)
        msg_text = (
            f"✏️ VCF Editor Mode\n"
            f"━━━━━━━━━━━━━━━\n"
            f"📁 Files Uploaded: {state['files']}\n"
            f"📊 Contacts Found: {len(state['contacts'])}\n"
            f"⏳ Status: Extracting...\n\n"
            f"📂 Keep sending files\n"
            f"✅ Finish Type → /done"
        )

        with msg_lock:
            if not state.get("msg_id"):
                msg = bot.send_message(message.chat.id, msg_text)
                state["msg_id"] = msg.message_id
            else:
                try:
                    bot.edit_message_text(msg_text, message.chat.id, state["msg_id"])
                except:
                    pass

        return


    elif filename.endswith(".vcf") and mode == "vcf_details":

        state["contacts"] = []
        state["file_name"] = filename
    
        with open(path, encoding="utf-8", errors="ignore") as f:
            current_name = ""
            current_phone = ""

            for line in f:
                line = line.strip()

                if line.startswith("FN:"):
                    current_name = line.replace("FN:", "").strip()

                elif "TEL" in line.upper():
                    num = line.split(":")[-1].strip()
                    num = num.replace(" ", "").replace("-", "")
                    if not num.startswith("+"):
                        num = "+" + num
                    current_phone = num

                elif "END:VCARD" in line:
                    if current_name or current_phone:
                        state["contacts"].append((current_name, current_phone))
                    current_name = ""
                    current_phone = ""

        os.remove(path)

        state["page"] = 1

        show_vcf_page(message.chat.id, state)
        send_txt_report(message.chat.id, state)
        return

    # ===== ADMIN/NAVY VCF SUPPORT =====
    if filename.endswith(".vcf") and mode == "admin_navy":

        contacts = []
        
        with open(path, encoding="utf-8", errors="ignore") as f:
            for line in f:
                if "TEL" in line.upper():
                    num = line.split(":")[-1].strip()
                    num = num.replace(" ", "").replace("-", "").replace("+", "")
                    if num.isdigit() and len(num) >= 8:
                        contacts.append(num)

        os.remove(path)

        # 👉 ADMIN STEP
        if state["step"] == "admin_collect":
            state["admin"].extend(contacts)
            update_admin_navy_msg(message, state, "admin")

        # 👉 NAVY STEP
        elif state["step"] == "navy_collect":
            state["navy"].extend(contacts)
            update_admin_navy_msg(message, state, "navy")

        return

        # ===== ADMIN/NAVY TXT SUPPORT =====
    if filename.endswith(".txt") and mode == "admin_navy":

        numbers = []

        with open(path, encoding="utf-8", errors="ignore") as f:
            for line in f:
                nums = extract_valid_numbers(line)
                numbers.extend(nums)

        os.remove(path)

        # 👉 ADMIN STEP
        if state["step"] == "admin_collect":
            state["admin"].extend(numbers)
            update_admin_navy_msg(message, state, "admin")

        # 👉 NAVY STEP
        elif state["step"] == "navy_collect":
            state["navy"].extend(numbers)
            update_admin_navy_msg(message, state, "navy")

        return

        # ===== ADMIN/NAVY XLSX SUPPORT =====
    if filename.endswith(".xlsx") and mode == "admin_navy":

        from openpyxl import load_workbook

        numbers = []

        wb = load_workbook(path, read_only=True)

        for row in wb.active.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    nums = extract_valid_numbers(str(cell))
                    numbers.extend(nums)

        wb.close()
        os.remove(path)

        # 👉 ADMIN STEP
        if state["step"] == "admin_collect":
            state["admin"].extend(numbers)
            update_admin_navy_msg(message, state, "admin")

        # 👉 NAVY STEP
        elif state["step"] == "navy_collect":
            state["navy"].extend(numbers)
            update_admin_navy_msg(message, state, "navy")

        return


    # ============================================================
    # INVALID
    # ============================================================
    os.remove(path)
    bot.send_message(message.chat.id, "❌ Invalid file type for current mode.")


# ============================================================
# 🔹 SHOW VCF PAGE
# ============================================================
def show_vcf_page(chat_id, state):
    contacts = state["contacts"]
    page = state["page"]
    per_page = 10

    total = len(contacts)
    total_pages = (total // per_page) + (1 if total % per_page else 0)

    start = (page - 1) * per_page
    end = start + per_page
    chunk = contacts[start:end]

    text = f"""╭━━━ 📁 VCF DETAILS ━━━╮
┃ 📄 File    : {state.get('file_name','Unknown')}
┃ 👥 Contacts: {total}
┃ 📊 Showing : {start+1}-{min(end, total)}
╰━━━━━━━━━━━━━━━━━━━━━━╯

┏━━━━━━━━━━━━━━━━━━━━━━┓
┃ 📄 Page {page} / {total_pages}
┗━━━━━━━━━━━━━━━━━━━━━━┛
"""

    for i, (name, phone) in enumerate(chunk, start=start+1):
        text += f"\n{i}. 👤 {name}\n   ┗ 📞 {phone}\n"

    kb = types.InlineKeyboardMarkup(row_width=2)

    buttons = []
    if page > 1:
        buttons.append(types.InlineKeyboardButton("⬅️ Prev", callback_data="prev"))
    if page < total_pages:
        buttons.append(types.InlineKeyboardButton("Next ➡️", callback_data="next"))

    if buttons:
        kb.row(*buttons)

    try:
        bot.edit_message_text(text, chat_id, state["msg_id"], reply_markup=kb)
    except:
        pass


# ============================================================
# 🔹 Run Bot
# ============================================================
def run_bot():
    print("🚀 Bot started")

    bot.remove_webhook()
    time.sleep(1)

    bot.infinity_polling(skip_pending=True, none_stop=True)

if __name__ == "__main__":
    threading.Thread(target=run_bot, daemon=True).start()

    port = int(os.getenv("PORT", 5000))
    web.run(host="0.0.0.0", port=port)